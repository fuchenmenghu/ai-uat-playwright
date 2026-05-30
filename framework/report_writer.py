from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from framework.ai_suggestion import generate_ai_suggestion
from framework.context import RunContext
from framework.step import RESULT_FAIL, RESULT_PASS


SUMMARY_HEADERS = [
    "run_id",
    "脚本编号",
    "脚本名称",
    "源头单号",
    "测试环境",
    "涉及系统",
    "开始时间",
    "结束时间",
    "总步骤数",
    "通过数",
    "不通过数",
    "最终测试结论",
    "首个失败步骤",
    "AI改善建议 / 潜在风险点",
]

OPERATION_HEADERS = [
    "脚本编号",
    "步骤编号",
    "系统",
    "模块",
    "操作描述",
    "预期结果",
    "执行结果",
    "失败编码",
    "失败类型",
    "原始失败信息",
    "截图路径",
    "trace路径",
    "执行时间",
]


def write_report(ctx: RunContext, reports_dir: Path) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    report_path = reports_dir / f"{ctx.run_id}_{ctx.script_id}_Test_Report.xlsx"

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "summary_report"
    operation_ws = wb.create_sheet("operation_log")

    _write_summary(summary_ws, ctx)
    _write_operation_log(operation_ws, ctx)

    wb.save(report_path)
    ctx.report_path = report_path
    return report_path


def _write_summary(ws, ctx: RunContext) -> None:
    total = len(ctx.logs)
    passed = sum(1 for item in ctx.logs if item.result == RESULT_PASS)
    failed = sum(1 for item in ctx.logs if item.result == RESULT_FAIL)
    conclusion = RESULT_FAIL if failed else RESULT_PASS
    first_failed = next((item.step_no for item in ctx.logs if item.result == RESULT_FAIL), "")

    started_at = ctx.started_at.strftime("%Y-%m-%d %H:%M:%S")
    ended_at = (ctx.ended_at or ctx.started_at).strftime("%Y-%m-%d %H:%M:%S")
    ai_suggestion = generate_ai_suggestion(ctx.logs)

    row = [
        ctx.run_id,
        ctx.script_id,
        ctx.script_name,
        ctx.source_order_no,
        ctx.env,
        ctx.involved_system_names,
        started_at,
        ended_at,
        total,
        passed,
        failed,
        conclusion,
        first_failed,
        ai_suggestion,
    ]
    _write_table(ws, SUMMARY_HEADERS, [row])


def _write_operation_log(ws, ctx: RunContext) -> None:
    rows = []
    for item in ctx.logs:
        row_dict = item.to_operation_log_row()
        rows.append([row_dict.get(header, "") for header in OPERATION_HEADERS])
    _write_table(ws, OPERATION_HEADERS, rows)


def _write_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column in ws.columns:
        max_length = 10
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
